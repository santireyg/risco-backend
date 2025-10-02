"""
Servicio de exportación de datos financieros a formato Excel (.xlsx)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from bson import ObjectId
from fastapi import HTTPException
from app.core.database import docs_collection
from app.models.users import User
import re
import logging
from io import BytesIO
import gc

logger = logging.getLogger(__name__)

# Constantes de color
COLOR_BORDER = '4a568d'
COLOR_HEADER_BG = '4a568d'
COLOR_HEADER_TEXT = 'FFFFFF'
COLOR_ODD_ROW = 'f6f8f9'
COLOR_EVEN_ROW = 'FFFFFF'
COLOR_TEXT = '48484e'
COLOR_TITLE = '4a568d'

# Constantes de altura de filas
ROW_HEIGHT_NORMAL = 23
ROW_HEIGHT_HEADER = 26

# Constantes de estilo
HEADER_FONT = Font(name='Calibri', size=12, bold=False, color=COLOR_HEADER_TEXT)
HEADER_FILL = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=False)

TITLE_FONT = Font(name='Calibri', size=14, bold=False, color=COLOR_TITLE)
LABEL_FONT = Font(name='Calibri', size=12, bold=True, color=COLOR_TEXT)
NORMAL_FONT = Font(name='Calibri', size=12, color=COLOR_TEXT)

DATA_FONT = Font(name='Calibri', size=11, color=COLOR_TEXT)
DATA_FONT_RED = Font(name='Calibri', size=11, color='FF0000')
DATA_ALIGNMENT_RIGHT = Alignment(horizontal='right', vertical='center')
DATA_ALIGNMENT_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

EVEN_ROW_FILL = PatternFill(start_color=COLOR_EVEN_ROW, end_color=COLOR_EVEN_ROW, fill_type='solid')
ODD_ROW_FILL = PatternFill(start_color=COLOR_ODD_ROW, end_color=COLOR_ODD_ROW, fill_type='solid')

# Bordes personalizados
BORDER_COLOR = Side(style='thin', color=COLOR_BORDER)
BORDER_NONE = Side(style=None)

# Bordes para encabezado (solo borde inferior)
HEADER_BORDER = Border(
    left=BORDER_NONE,
    right=BORDER_NONE,
    top=BORDER_NONE,
    bottom=BORDER_COLOR
)

# Bordes para celdas de datos (sin bordes internos)
DATA_BORDER_NONE = Border(
    left=BORDER_NONE,
    right=BORDER_NONE,
    top=BORDER_NONE,
    bottom=BORDER_NONE
)

# Formato de número contabilidad
ACCOUNTING_FORMAT = '#,##0.00;[Red](#,##0.00)'


def sanitize_filename(text: str, max_length: int = 100) -> str:
    """
    Sanitiza un texto para usarlo como nombre de archivo.
    Remueve caracteres no válidos y limita la longitud.
    """
    # Remover caracteres no válidos
    invalid_chars = r'[/\\:*?"<>|]'
    text = re.sub(invalid_chars, '', text)
    
    # Reemplazar espacios por guiones bajos
    text = text.replace(' ', '_')
    
    # Limitar longitud
    if len(text) > max_length:
        text = text[:max_length]
    
    return text


def format_date(dt: datetime) -> str:
    """Formatea un datetime en formato DD/MM/YYYY"""
    if dt is None:
        return "No disponible"
    return dt.strftime("%d/%m/%Y")


def format_datetime(dt: datetime) -> str:
    """Formatea un datetime en formato DD/MM/YYYY HH:MM"""
    return dt.strftime("%d/%m/%Y %H:%M")


def format_cuit(cuit: str) -> str:
    """
    Formatea un CUIT en formato XX-XXXXXXXX-X
    """
    if not cuit or cuit == "No disponible":
        return "No disponible"
    
    # Remover cualquier caracter que no sea número
    cuit_numbers = re.sub(r'\D', '', cuit)
    
    # Verificar que tenga 11 dígitos
    if len(cuit_numbers) != 11:
        return cuit  # Retornar el original si no tiene el formato esperado
    
    # Aplicar formato XX-XXXXXXXX-X
    return f"{cuit_numbers[:2]}-{cuit_numbers[2:10]}-{cuit_numbers[10]}"


def add_header_info(ws, company_info: dict, general_info: dict, current_row: int = 1) -> int:
    """
    Agrega el encabezado de información general a una hoja.
    Retorna la siguiente fila disponible.
    """
    # CUIT
    ws[f'A{current_row}'] = "CUIT:"
    ws[f'A{current_row}'].font = LABEL_FONT
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    cuit_value = company_info.get('company_cuit')
    ws[f'B{current_row}'] = format_cuit(cuit_value) if cuit_value else "No disponible"
    ws[f'B{current_row}'].font = NORMAL_FONT
    ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    current_row += 1
    
    # Razón social
    ws[f'A{current_row}'] = "Razón social:"
    ws[f'A{current_row}'].font = LABEL_FONT
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'B{current_row}'] = company_info.get('company_name', '')
    ws[f'B{current_row}'].font = NORMAL_FONT
    ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    current_row += 1
    
    # Período actual
    ws[f'A{current_row}'] = "Período actual:"
    ws[f'A{current_row}'].font = LABEL_FONT
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    periodo_actual = general_info.get('periodo_actual')
    ws[f'B{current_row}'] = format_date(periodo_actual)
    ws[f'B{current_row}'].font = NORMAL_FONT
    ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    current_row += 1
    
    # Período anterior
    ws[f'A{current_row}'] = "Período anterior:"
    ws[f'A{current_row}'].font = LABEL_FONT
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    periodo_anterior = general_info.get('periodo_anterior')
    ws[f'B{current_row}'] = format_date(periodo_anterior)
    ws[f'B{current_row}'].font = NORMAL_FONT
    ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    current_row += 1
    
    # Fecha y hora de exportación
    ws[f'A{current_row}'] = "Fecha y hora exportación:"
    ws[f'A{current_row}'].font = LABEL_FONT
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'B{current_row}'] = format_datetime(datetime.now())
    ws[f'B{current_row}'].font = NORMAL_FONT
    ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    current_row += 1
    
    # Separación (2 filas vacías)
    return current_row + 2


def add_table(ws, title: str, data: list, current_row: int) -> int:
    """
    Agrega una tabla con formato a la hoja.
    data debe ser una lista de dicts con keys: concepto, monto_actual, monto_anterior
    Retorna la siguiente fila disponible.
    """
    # Título de la tabla
    ws[f'A{current_row}'] = title
    ws[f'A{current_row}'].font = TITLE_FONT
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    current_row += 1
    
    # Guardar la fila del encabezado para aplicar bordes después
    header_row = current_row
    
    # Encabezado de columnas
    ws[f'A{current_row}'] = "Concepto"
    ws[f'B{current_row}'] = "Actual"
    ws[f'C{current_row}'] = "Anterior"
    
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{current_row}']
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER
    
    # Establecer altura de la fila del encabezado
    ws.row_dimensions[current_row].height = ROW_HEIGHT_HEADER
    
    current_row += 1
    first_data_row = current_row
    
    # Filas de datos
    if not data or len(data) == 0:
        # Sin datos disponibles
        ws[f'A{current_row}'] = "Sin datos disponibles"
        ws[f'A{current_row}'].font = DATA_FONT
        ws[f'A{current_row}'].alignment = DATA_ALIGNMENT_LEFT
        ws[f'A{current_row}'].border = DATA_BORDER_NONE
        ws[f'A{current_row}'].fill = EVEN_ROW_FILL
        
        ws[f'B{current_row}'].border = DATA_BORDER_NONE
        ws[f'B{current_row}'].fill = EVEN_ROW_FILL
        
        ws[f'C{current_row}'].border = DATA_BORDER_NONE
        ws[f'C{current_row}'].fill = EVEN_ROW_FILL
        
        # Establecer altura de fila
        ws.row_dimensions[current_row].height = ROW_HEIGHT_NORMAL
        
        current_row += 1
    else:
        for idx, item in enumerate(data):
            # Alternar colores de fondo
            fill = EVEN_ROW_FILL if idx % 2 == 0 else ODD_ROW_FILL
            
            # Concepto
            concepto = item.get('concepto', item.get('concepto_code', ''))
            ws[f'A{current_row}'] = concepto
            ws[f'A{current_row}'].font = DATA_FONT
            ws[f'A{current_row}'].alignment = DATA_ALIGNMENT_LEFT
            ws[f'A{current_row}'].border = DATA_BORDER_NONE
            ws[f'A{current_row}'].fill = fill
            
            # Monto actual
            monto_actual = item.get('monto_actual', 0)
            if monto_actual != 0:
                ws[f'B{current_row}'] = monto_actual
                ws[f'B{current_row}'].number_format = ACCOUNTING_FORMAT
                # Aplicar color rojo si es negativo
                if monto_actual < 0:
                    ws[f'B{current_row}'].font = DATA_FONT_RED
                else:
                    ws[f'B{current_row}'].font = DATA_FONT
            else:
                ws[f'B{current_row}'] = None
                ws[f'B{current_row}'].font = DATA_FONT
            ws[f'B{current_row}'].alignment = DATA_ALIGNMENT_RIGHT
            ws[f'B{current_row}'].border = DATA_BORDER_NONE
            ws[f'B{current_row}'].fill = fill
            
            # Monto anterior
            monto_anterior = item.get('monto_anterior', 0)
            if monto_anterior != 0:
                ws[f'C{current_row}'] = monto_anterior
                ws[f'C{current_row}'].number_format = ACCOUNTING_FORMAT
                # Aplicar color rojo si es negativo
                if monto_anterior < 0:
                    ws[f'C{current_row}'].font = DATA_FONT_RED
                else:
                    ws[f'C{current_row}'].font = DATA_FONT
            else:
                ws[f'C{current_row}'] = None
                ws[f'C{current_row}'].font = DATA_FONT
            ws[f'C{current_row}'].alignment = DATA_ALIGNMENT_RIGHT
            ws[f'C{current_row}'].border = DATA_BORDER_NONE
            ws[f'C{current_row}'].fill = fill
            
            # Establecer altura de fila
            ws.row_dimensions[current_row].height = ROW_HEIGHT_NORMAL
            
            current_row += 1
    
    last_data_row = current_row - 1
    
    # Aplicar bordes externos a toda la tabla (desde encabezado hasta última fila de datos)
    # Borde izquierdo (columna A)
    for row in range(header_row, last_data_row + 1):
        cell = ws[f'A{row}']
        cell.border = Border(
            left=BORDER_COLOR,
            right=cell.border.right if cell.border else BORDER_NONE,
            top=cell.border.top if cell.border else BORDER_NONE,
            bottom=cell.border.bottom if cell.border else BORDER_NONE
        )
    
    # Borde derecho (columna C)
    for row in range(header_row, last_data_row + 1):
        cell = ws[f'C{row}']
        cell.border = Border(
            left=cell.border.left if cell.border else BORDER_NONE,
            right=BORDER_COLOR,
            top=cell.border.top if cell.border else BORDER_NONE,
            bottom=cell.border.bottom if cell.border else BORDER_NONE
        )
    
    # Borde superior (fila del encabezado)
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{header_row}']
        cell.border = Border(
            left=cell.border.left if cell.border else BORDER_NONE,
            right=cell.border.right if cell.border else BORDER_NONE,
            top=BORDER_COLOR,
            bottom=cell.border.bottom if cell.border else BORDER_NONE
        )
    
    # Borde inferior (última fila de datos)
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{last_data_row}']
        cell.border = Border(
            left=cell.border.left if cell.border else BORDER_NONE,
            right=cell.border.right if cell.border else BORDER_NONE,
            top=cell.border.top if cell.border else BORDER_NONE,
            bottom=BORDER_COLOR
        )
    
    # Separación (2 filas vacías)
    return current_row + 2


def configure_sheet(ws):
    """Configura los ajustes generales de una hoja"""
    # Desactivar líneas de cuadrícula
    ws.sheet_view.showGridLines = False
    
    # Anchos de columna
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25


def create_situacion_patrimonial_sheet(wb, document: dict) -> bool:
    """
    Crea la hoja de Situación Patrimonial.
    Retorna True si se creó exitosamente, False si no hay datos.
    """
    balance_data = document.get('balance_data')
    if not balance_data:
        return False
    
    ws = wb.active
    ws.title = "Situación Patrimonial"
    
    company_info = document.get('company_info', {})
    general_info = balance_data.get('informacion_general', {})
    
    # Agregar encabezado
    current_row = add_header_info(ws, company_info, general_info, 1)
    
    # Tabla Activo
    detalles_activo = balance_data.get('detalles_activo', [])
    current_row = add_table(ws, "Activo", detalles_activo, current_row)
    
    # Tabla Pasivo
    detalles_pasivo = balance_data.get('detalles_pasivo', [])
    current_row = add_table(ws, "Pasivo", detalles_pasivo, current_row)
    
    # Tabla Patrimonio Neto
    detalles_patrimonio_neto = balance_data.get('detalles_patrimonio_neto', [])
    current_row = add_table(ws, "Patrimonio Neto", detalles_patrimonio_neto, current_row)
    
    configure_sheet(ws)
    return True


def create_estado_resultados_sheet(wb, document: dict) -> bool:
    """
    Crea la hoja de Estado de Resultados.
    Retorna True si se creó exitosamente, False si no hay datos.
    """
    income_data = document.get('income_statement_data')
    if not income_data:
        return False
    
    ws = wb.create_sheet("Estado Resultados")
    
    company_info = document.get('company_info', {})
    general_info = income_data.get('informacion_general', {})
    
    # Agregar encabezado
    current_row = add_header_info(ws, company_info, general_info, 1)
    
    # Tabla Estado de Resultados
    detalles_estado_resultados = income_data.get('detalles_estado_resultados', [])
    current_row = add_table(ws, "Estado de Resultados", detalles_estado_resultados, current_row)
    
    configure_sheet(ws)
    return True


def create_cuentas_principales_sheet(wb, document: dict) -> bool:
    """
    Crea la hoja de Cuentas Principales.
    Retorna True si se creó exitosamente, False si no hay datos.
    """
    balance_data = document.get('balance_data')
    income_data = document.get('income_statement_data')
    
    if not balance_data and not income_data:
        return False
    
    ws = wb.create_sheet("Cuentas Principales")
    
    company_info = document.get('company_info', {})
    
    # Usar información general del balance o income (el que esté disponible)
    general_info = {}
    if balance_data:
        general_info = balance_data.get('informacion_general', {})
    elif income_data:
        general_info = income_data.get('informacion_general', {})
    
    # Agregar encabezado
    current_row = add_header_info(ws, company_info, general_info, 1)
    
    # Título
    ws[f'A{current_row}'] = "Cuentas Principales"
    ws[f'A{current_row}'].font = TITLE_FONT
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    current_row += 1
    
    # Guardar la fila del encabezado
    header_row = current_row
    
    # Encabezado de columnas
    ws[f'A{current_row}'] = "Concepto"
    ws[f'B{current_row}'] = "Actual"
    ws[f'C{current_row}'] = "Anterior"
    
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{current_row}']
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER
    
    # Establecer altura de la fila del encabezado
    ws.row_dimensions[current_row].height = ROW_HEIGHT_HEADER
    
    current_row += 1
    first_data_row = current_row
    
    # Combinar datos de balance y income
    combined_data = []
    
    if balance_data:
        resultados_principales_balance = balance_data.get('resultados_principales', [])
        combined_data.extend(resultados_principales_balance)
    
    if income_data:
        resultados_principales_income = income_data.get('resultados_principales', [])
        combined_data.extend(resultados_principales_income)
    
    # Agregar filas de datos
    if not combined_data or len(combined_data) == 0:
        # Sin datos disponibles
        ws[f'A{current_row}'] = "Sin datos disponibles"
        ws[f'A{current_row}'].font = DATA_FONT
        ws[f'A{current_row}'].alignment = DATA_ALIGNMENT_LEFT
        ws[f'A{current_row}'].border = DATA_BORDER_NONE
        ws[f'A{current_row}'].fill = EVEN_ROW_FILL
        
        ws[f'B{current_row}'].border = DATA_BORDER_NONE
        ws[f'B{current_row}'].fill = EVEN_ROW_FILL
        
        ws[f'C{current_row}'].border = DATA_BORDER_NONE
        ws[f'C{current_row}'].fill = EVEN_ROW_FILL
        
        # Establecer altura de fila
        ws.row_dimensions[current_row].height = ROW_HEIGHT_NORMAL
        
        current_row += 1
    else:
        for idx, item in enumerate(combined_data):
            # Alternar colores de fondo
            fill = EVEN_ROW_FILL if idx % 2 == 0 else ODD_ROW_FILL
            
            # Concepto
            concepto = item.get('concepto', item.get('concepto_code', ''))
            ws[f'A{current_row}'] = concepto
            ws[f'A{current_row}'].font = DATA_FONT
            ws[f'A{current_row}'].alignment = DATA_ALIGNMENT_LEFT
            ws[f'A{current_row}'].border = DATA_BORDER_NONE
            ws[f'A{current_row}'].fill = fill
            
            # Monto actual
            monto_actual = item.get('monto_actual', 0)
            if monto_actual != 0:
                ws[f'B{current_row}'] = monto_actual
                ws[f'B{current_row}'].number_format = ACCOUNTING_FORMAT
                # Aplicar color rojo si es negativo
                if monto_actual < 0:
                    ws[f'B{current_row}'].font = DATA_FONT_RED
                else:
                    ws[f'B{current_row}'].font = DATA_FONT
            else:
                ws[f'B{current_row}'] = None
                ws[f'B{current_row}'].font = DATA_FONT
            ws[f'B{current_row}'].alignment = DATA_ALIGNMENT_RIGHT
            ws[f'B{current_row}'].border = DATA_BORDER_NONE
            ws[f'B{current_row}'].fill = fill
            
            # Monto anterior
            monto_anterior = item.get('monto_anterior', 0)
            if monto_anterior != 0:
                ws[f'C{current_row}'] = monto_anterior
                ws[f'C{current_row}'].number_format = ACCOUNTING_FORMAT
                # Aplicar color rojo si es negativo
                if monto_anterior < 0:
                    ws[f'C{current_row}'].font = DATA_FONT_RED
                else:
                    ws[f'C{current_row}'].font = DATA_FONT
            else:
                ws[f'C{current_row}'] = None
                ws[f'C{current_row}'].font = DATA_FONT
            ws[f'C{current_row}'].alignment = DATA_ALIGNMENT_RIGHT
            ws[f'C{current_row}'].border = DATA_BORDER_NONE
            ws[f'C{current_row}'].fill = fill
            
            # Establecer altura de fila
            ws.row_dimensions[current_row].height = ROW_HEIGHT_NORMAL
            
            current_row += 1
    
    last_data_row = current_row - 1
    
    # Aplicar bordes externos a toda la tabla (desde encabezado hasta última fila de datos)
    # Borde izquierdo (columna A)
    for row in range(header_row, last_data_row + 1):
        cell = ws[f'A{row}']
        cell.border = Border(
            left=BORDER_COLOR,
            right=cell.border.right if cell.border else BORDER_NONE,
            top=cell.border.top if cell.border else BORDER_NONE,
            bottom=cell.border.bottom if cell.border else BORDER_NONE
        )
    
    # Borde derecho (columna C)
    for row in range(header_row, last_data_row + 1):
        cell = ws[f'C{row}']
        cell.border = Border(
            left=cell.border.left if cell.border else BORDER_NONE,
            right=BORDER_COLOR,
            top=cell.border.top if cell.border else BORDER_NONE,
            bottom=cell.border.bottom if cell.border else BORDER_NONE
        )
    
    # Borde superior (fila del encabezado)
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{header_row}']
        cell.border = Border(
            left=cell.border.left if cell.border else BORDER_NONE,
            right=cell.border.right if cell.border else BORDER_NONE,
            top=BORDER_COLOR,
            bottom=cell.border.bottom if cell.border else BORDER_NONE
        )
    
    # Borde inferior (última fila de datos)
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{last_data_row}']
        cell.border = Border(
            left=cell.border.left if cell.border else BORDER_NONE,
            right=cell.border.right if cell.border else BORDER_NONE,
            top=cell.border.top if cell.border else BORDER_NONE,
            bottom=BORDER_COLOR
        )
    
    configure_sheet(ws)
    return True


def generate_filename(document: dict) -> str:
    """
    Genera un nombre de archivo sanitizado para el documento.
    Formato: {company_name}_{periodo_actual}.xlsx
    """
    company_info = document.get('company_info', {})
    company_name = company_info.get('company_name')
    
    # Usar balance_data o income_statement_data para obtener el período
    periodo_actual = None
    balance_data = document.get('balance_data')
    income_data = document.get('income_statement_data')
    
    if balance_data:
        general_info = balance_data.get('informacion_general', {})
        periodo_actual = general_info.get('periodo_actual')
    elif income_data:
        general_info = income_data.get('informacion_general', {})
        periodo_actual = general_info.get('periodo_actual')
    
    # Si no hay company_name, usar el ID del documento
    if not company_name:
        company_name = f"documento_{document.get('_id')}"
    
    # Si no hay periodo_actual, usar fecha actual
    if not periodo_actual:
        periodo_str = datetime.now().strftime("%d_%m_%Y")
    else:
        periodo_str = periodo_actual.strftime("%d_%m_%Y")
    
    filename = f"{company_name}_{periodo_str}"
    filename = sanitize_filename(filename)
    filename += ".xlsx"
    
    return filename


async def generate_excel_export(docfile_id: str, current_user: User) -> bytes:
    """
    Genera un archivo Excel con los datos financieros del documento.
    
    Args:
        docfile_id: ID del documento en MongoDB
        current_user: Usuario actual autenticado
    
    Returns:
        bytes: Contenido del archivo Excel
    
    Raises:
        HTTPException: Si hay errores de validación o generación
    """
    try:
        # Validar ObjectId
        if not ObjectId.is_valid(docfile_id):
            raise HTTPException(status_code=400, detail="ID de documento inválido")
        
        # Obtener documento
        document = await docs_collection.find_one({"_id": ObjectId(docfile_id)})
        
        if not document:
            raise HTTPException(status_code=404, detail="Documento no encontrado")
        
        # Validar pertenencia al tenant
        if document.get('tenant_id') != current_user.tenant_id:
            raise HTTPException(
                status_code=403,
                detail="No tiene permisos para exportar este documento"
            )
        
        # Validar status
        status = document.get('status')
        if status not in ["Analizado", "Exportado"]:
            raise HTTPException(
                status_code=422,
                detail=f"El documento debe estar en estado 'Analizado' o 'Exportado' para ser exportado. Estado actual: {status}"
            )
        
        # Validar company_info
        company_info = document.get('company_info')
        if not company_info:
            raise HTTPException(
                status_code=422,
                detail="El documento debe tener información de la empresa para exportar"
            )
        
        # Validar que haya datos financieros
        balance_data = document.get('balance_data')
        income_data = document.get('income_statement_data')
        
        if not balance_data and not income_data:
            raise HTTPException(
                status_code=422,
                detail="El documento no tiene datos financieros para exportar"
            )
        
        # Crear workbook
        wb = Workbook()
        sheets_created = 0
        
        # Crear hoja de Situación Patrimonial
        if create_situacion_patrimonial_sheet(wb, document):
            sheets_created += 1
        else:
            # Si no se creó, eliminar la hoja activa por defecto
            if len(wb.sheetnames) == 1 and wb.active.title == "Sheet":
                wb.remove(wb.active)
        
        # Crear hoja de Estado de Resultados
        if create_estado_resultados_sheet(wb, document):
            sheets_created += 1
        
        # Crear hoja de Cuentas Principales
        if create_cuentas_principales_sheet(wb, document):
            sheets_created += 1
        
        if sheets_created == 0:
            raise HTTPException(
                status_code=422,
                detail="No se pudo generar ninguna hoja con los datos disponibles"
            )
        
        # Guardar en memoria
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        excel_bytes = buffer.getvalue()
        
        # Limpiar memoria
        buffer.close()
        del wb
        gc.collect()
        
        logger.info(
            f"Exportación Excel completada para documento {docfile_id}, "
            f"tamaño: {len(excel_bytes)} bytes"
        )
        
        return excel_bytes
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al exportar documento {docfile_id}: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error interno al generar el archivo Excel: {str(e)}"
        )
